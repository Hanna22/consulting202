"""Split the first worksheet of an Excel workbook into one file per key value.

Edit the values in the CONFIGURATION section below, or pass command-line
arguments.  The implementation uses a temporary SQLite database so even a
large source workbook does not need to be held entirely in memory.

Requires: openpyxl (``pip install openpyxl``)
"""

from __future__ import annotations

import argparse
import logging
import pickle
import re
import sqlite3
import sys
import tempfile
from contextlib import closing
from datetime import date, datetime, time
from pathlib import Path
from typing import Any, Iterator, Sequence

from openpyxl import Workbook, load_workbook


# ---------------------------------------------------------------------------
# CONFIGURATION - these values are intentionally easy to edit.
# Command-line arguments override them when supplied.
# ---------------------------------------------------------------------------
INPUT_FILE = Path(r"C:\Users\hkildani002\Downloads\210726_Consolidated_Catalog3.xlsx")
OUTPUT_DIRECTORY = Path(r"C:\Users\hkildani002\Downloads\split_excels_output")
KEY_COLUMN_NAME = "Entity"
OUTPUT_FILE_PREFIX = "ATTR_"
CASE_SENSITIVE_COLUMN_NAME = False
DATA_ONLY = True  # True uses cached formula results instead of formula text.


EXCEL_MAX_ROWS = 1_048_576
PROGRESS_INTERVAL = 10_000
INVALID_FILENAME_CHARACTERS = re.compile(r'[<>:"/\\|?*\x00-\x1f]')
WINDOWS_RESERVED_NAMES = {
    "CON", "PRN", "AUX", "NUL",
    *(f"COM{number}" for number in range(1, 10)),
    *(f"LPT{number}" for number in range(1, 10)),
}

LOGGER = logging.getLogger("split_excels")


def configure_logging() -> None:
    """Configure intentionally verbose console logging."""
    logging.basicConfig(
        level=logging.DEBUG,
        format="%(asctime)s | %(levelname)-8s | %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
    )


def parse_arguments() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Split the first sheet of an .xlsx/.xlsm workbook into separate "
            "workbooks according to a key column."
        )
    )
    parser.add_argument(
        "input_file",
        nargs="?",
        type=Path,
        default=INPUT_FILE,
        help=f"Source workbook (default: {INPUT_FILE})",
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=OUTPUT_DIRECTORY,
        help=f"Directory for split files (default: {OUTPUT_DIRECTORY})",
    )
    parser.add_argument(
        "--key-column",
        default=KEY_COLUMN_NAME,
        help=f"Column header used for splitting (default: {KEY_COLUMN_NAME!r})",
    )
    parser.add_argument(
        "--prefix",
        default=OUTPUT_FILE_PREFIX,
        help=f"Output filename prefix (default: {OUTPUT_FILE_PREFIX!r})",
    )
    parser.add_argument(
        "--case-sensitive-column",
        action="store_true",
        default=CASE_SENSITIVE_COLUMN_NAME,
        help="Match the key column header with case sensitivity.",
    )
    parser.add_argument(
        "--formulas",
        action="store_true",
        help="Read formula text instead of cached formula results.",
    )
    return parser.parse_args()


def validate_settings(input_file: Path, output_directory: Path, key_column: str) -> None:
    """Fail early with useful messages for common configuration errors."""
    if not key_column.strip():
        raise ValueError("KEY_COLUMN_NAME / --key-column cannot be blank.")
    if not input_file.exists():
        raise FileNotFoundError(f"Input workbook does not exist: {input_file.resolve()}")
    if not input_file.is_file():
        raise ValueError(f"Input path is not a file: {input_file.resolve()}")
    if input_file.suffix.lower() not in {".xlsx", ".xlsm"}:
        raise ValueError(
            "This program supports .xlsx and .xlsm files. "
            f"Received: {input_file.suffix or '(no extension)'}"
        )

    if output_directory.exists() and not output_directory.is_dir():
        raise ValueError(f"Output path exists but is not a directory: {output_directory.resolve()}")
    output_directory.mkdir(parents=True, exist_ok=True)
    if output_directory.resolve() == input_file.resolve():
        raise ValueError("The output directory cannot be the input workbook.")


def find_key_column(headers: Sequence[Any], requested_name: str, case_sensitive: bool) -> int:
    """Return the zero-based index of the requested header."""
    requested = requested_name.strip()
    matches: list[int] = []

    for index, header in enumerate(headers):
        header_text = "" if header is None else str(header).strip()
        is_match = (
            header_text == requested
            if case_sensitive
            else header_text.casefold() == requested.casefold()
        )
        if is_match:
            matches.append(index)

    if not matches:
        available = ", ".join(repr(header) for header in headers)
        raise KeyError(
            f"Key column {requested_name!r} was not found in the first row. "
            f"Available headers: {available}"
        )
    if len(matches) > 1:
        raise ValueError(
            f"Key column {requested_name!r} matched more than one column "
            f"(Excel columns: {', '.join(str(index + 1) for index in matches)})."
        )
    return matches[0]


def normalize_row_width(row: Sequence[Any], column_count: int) -> list[Any]:
    """Return exactly one value per master-file column, preserving empty columns."""
    values = list(row[:column_count])
    if len(values) < column_count:
        values.extend([None] * (column_count - len(values)))
    return values


def key_identity(value: Any) -> str:
    """Create a stable group identity while distinguishing unlike value types."""
    if value is None or (isinstance(value, str) and not value.strip()):
        return "blank:"
    if isinstance(value, (datetime, date, time)):
        return f"{type(value).__name__}:{value.isoformat()}"
    return f"{type(value).__name__}:{value!r}"


def display_key(value: Any) -> str:
    if value is None or (isinstance(value, str) and not value.strip()):
        return "blank"
    if isinstance(value, (datetime, date, time)):
        return value.isoformat().replace(":", "-")
    return str(value).strip()


def safe_filename_component(value: str, fallback: str = "blank") -> str:
    """Convert arbitrary Excel content into a safe, reasonably short filename."""
    cleaned = INVALID_FILENAME_CHARACTERS.sub("_", value)
    cleaned = re.sub(r"\s+", " ", cleaned).strip(" .")
    cleaned = cleaned or fallback
    if cleaned.upper() in WINDOWS_RESERVED_NAMES:
        cleaned = f"_{cleaned}"
    return cleaned[:120].rstrip(" .") or fallback


def choose_unique_stem(prefix: str, key_text: str, used_stems: set[str]) -> str:
    """Choose a case-insensitively unique output stem."""
    base = safe_filename_component(prefix, fallback="") + safe_filename_component(key_text)
    base = base[:120].rstrip(" .") or "split_blank"
    candidate = base
    counter = 2
    while candidate.casefold() in used_stems:
        suffix = f"_{counter}"
        candidate = f"{base[:120 - len(suffix)]}{suffix}"
        counter += 1
    used_stems.add(candidate.casefold())
    return candidate


def stage_source_rows(
    input_file: Path,
    database: sqlite3.Connection,
    key_column_name: str,
    case_sensitive: bool,
    data_only: bool,
) -> tuple[list[Any], str, int, int]:
    """Stream the first sheet into SQLite and return basic source metadata."""
    LOGGER.info("Opening source workbook in read-only mode: %s", input_file.resolve())
    workbook = load_workbook(input_file, read_only=True, data_only=data_only)
    try:
        if not workbook.sheetnames:
            raise ValueError("The input workbook contains no worksheets.")

        first_sheet_name = workbook.sheetnames[0]
        ignored_sheets = workbook.sheetnames[1:]
        worksheet = workbook[first_sheet_name]
        LOGGER.info("Using first worksheet: %r", first_sheet_name)
        if ignored_sheets:
            LOGGER.info("Ignoring %d additional worksheet(s): %s", len(ignored_sheets), ignored_sheets)

        row_iterator = worksheet.iter_rows(values_only=True)
        try:
            headers = list(next(row_iterator))
        except StopIteration as error:
            raise ValueError(f"The first worksheet {first_sheet_name!r} is empty.") from error

        if not headers or all(header is None for header in headers):
            raise ValueError("The first row must contain column headers.")

        key_index = find_key_column(headers, key_column_name, case_sensitive)
        LOGGER.info(
            "Splitting on header %r in Excel column %d.",
            headers[key_index],
            key_index + 1,
        )
        LOGGER.info(
            "Master column structure contains %s column(s); every output file will retain all of them in this order.",
            f"{len(headers):,}",
        )

        database.executescript(
            """
            CREATE TABLE groups (
                key_id TEXT PRIMARY KEY,
                display_value TEXT NOT NULL,
                row_count INTEGER NOT NULL DEFAULT 0
            );
            CREATE TABLE rows (
                key_id TEXT NOT NULL,
                source_row INTEGER NOT NULL,
                payload BLOB NOT NULL
            );
            CREATE INDEX rows_key_id ON rows(key_id, source_row);
            """
        )

        total_rows = 0
        blank_key_rows = 0
        for source_row, row in enumerate(row_iterator, start=2):
            # A group may have no values at all in one or more columns. Always
            # pad each row to the complete master header width so those columns
            # remain present and in their original positions in every output.
            row_values = normalize_row_width(row, len(headers))

            key_value = row_values[key_index]
            identity = key_identity(key_value)
            shown_value = display_key(key_value)
            if identity == "blank:":
                blank_key_rows += 1

            database.execute(
                "INSERT INTO groups(key_id, display_value, row_count) VALUES (?, ?, 1) "
                "ON CONFLICT(key_id) DO UPDATE SET row_count = row_count + 1",
                (identity, shown_value),
            )
            database.execute(
                "INSERT INTO rows(key_id, source_row, payload) VALUES (?, ?, ?)",
                (identity, source_row, pickle.dumps(row_values, protocol=pickle.HIGHEST_PROTOCOL)),
            )
            total_rows += 1

            if total_rows % PROGRESS_INTERVAL == 0:
                database.commit()
                LOGGER.info("Staged %s data rows so far...", f"{total_rows:,}")

        database.commit()
        group_count = database.execute("SELECT COUNT(*) FROM groups").fetchone()[0]
        LOGGER.info(
            "Finished reading %s data rows across %s distinct key value(s).",
            f"{total_rows:,}",
            f"{group_count:,}",
        )
        if blank_key_rows:
            LOGGER.warning(
                "%s row(s) have a blank key; they will be written to the 'blank' group.",
                f"{blank_key_rows:,}",
            )
        if total_rows == 0:
            LOGGER.warning("The first worksheet has headers but contains no data rows.")
        return headers, first_sheet_name, total_rows, group_count
    finally:
        workbook.close()
        LOGGER.debug("Closed source workbook.")


def iter_group_rows(database: sqlite3.Connection, key_id: str) -> Iterator[list[Any]]:
    cursor = database.execute(
        "SELECT payload FROM rows WHERE key_id = ? ORDER BY source_row",
        (key_id,),
    )
    for (payload,) in cursor:
        yield pickle.loads(payload)


def write_group_workbook(
    output_path: Path,
    headers: Sequence[Any],
    rows: Iterator[list[Any]],
    expected_rows: int,
) -> int:
    """Write one group, rolling over to another sheet at Excel's row limit."""
    workbook = Workbook(write_only=True)
    sheet_number = 1
    worksheet = workbook.create_sheet(title="Data")
    worksheet.append(list(headers))
    rows_on_sheet = 0
    written = 0

    for row in rows:
        if rows_on_sheet >= EXCEL_MAX_ROWS - 1:
            sheet_number += 1
            worksheet = workbook.create_sheet(title=f"Data_{sheet_number}")
            worksheet.append(list(headers))
            rows_on_sheet = 0
            LOGGER.warning(
                "Excel row limit reached for %s; continuing on worksheet %s.",
                output_path.name,
                worksheet.title,
            )
        normalized_row = normalize_row_width(row, len(headers))
        if len(normalized_row) != len(headers):
            raise RuntimeError(
                f"Internal column-count mismatch for {output_path.name}: "
                f"expected {len(headers):,} columns."
            )
        worksheet.append(normalized_row)
        rows_on_sheet += 1
        written += 1

    if written != expected_rows:
        raise RuntimeError(
            f"Internal row-count mismatch for {output_path.name}: "
            f"expected {expected_rows:,}, wrote {written:,}."
        )
    workbook.save(output_path)
    workbook.close()
    return sheet_number


def verify_output_column_structure(output_path: Path, expected_headers: Sequence[Any]) -> None:
    """Confirm every output sheet has the full master header structure."""
    workbook = load_workbook(output_path, read_only=True, data_only=False)
    try:
        expected = tuple(expected_headers)
        for worksheet in workbook.worksheets:
            actual = tuple(
                next(
                    worksheet.iter_rows(
                        min_row=1,
                        max_row=1,
                        min_col=1,
                        max_col=len(expected),
                        values_only=True,
                    )
                )
            )
            if actual != expected:
                raise RuntimeError(
                    f"Column-structure verification failed for {output_path.name}, "
                    f"worksheet {worksheet.title!r}."
                )
    finally:
        workbook.close()


def create_split_files(
    database: sqlite3.Connection,
    headers: Sequence[Any],
    output_directory: Path,
    prefix: str,
) -> tuple[int, int]:
    groups = database.execute(
        "SELECT key_id, display_value, row_count FROM groups ORDER BY display_value, key_id"
    ).fetchall()
    # Do not silently overwrite workbooks from an earlier run. Existing names
    # are reserved, so a safe numeric suffix will be added when needed.
    used_stems = {
        path.stem.casefold()
        for path in output_directory.iterdir()
        if path.is_file() and path.suffix.casefold() == ".xlsx"
    }
    if used_stems:
        LOGGER.warning(
            "Output directory already contains %s .xlsx file(s); existing files will not be overwritten.",
            f"{len(used_stems):,}",
        )
    total_written = 0

    for group_number, (key_id, shown_value, row_count) in enumerate(groups, start=1):
        stem = choose_unique_stem(prefix, shown_value, used_stems)
        output_path = output_directory / f"{stem}.xlsx"
        LOGGER.info(
            "[%d/%d] Writing %s row(s) for key %r to %s",
            group_number,
            len(groups),
            f"{row_count:,}",
            shown_value,
            output_path.resolve(),
        )
        sheet_count = write_group_workbook(
            output_path,
            headers,
            iter_group_rows(database, key_id),
            row_count,
        )
        verify_output_column_structure(output_path, headers)
        total_written += row_count
        LOGGER.info(
            "[%d/%d] Completed %s (%d worksheet%s); verified all %s master column(s).",
            group_number,
            len(groups),
            output_path.name,
            sheet_count,
            "" if sheet_count == 1 else "s",
            f"{len(headers):,}",
        )

    return len(groups), total_written


def split_excel(
    input_file: Path,
    output_directory: Path,
    key_column: str,
    prefix: str,
    case_sensitive: bool = False,
    data_only: bool = True,
) -> None:
    """Run the complete split operation."""
    input_file = input_file.expanduser()
    output_directory = output_directory.expanduser()
    validate_settings(input_file, output_directory, key_column)

    LOGGER.info("Starting Excel split operation.")
    LOGGER.info("Input workbook : %s", input_file.resolve())
    LOGGER.info("Output directory: %s", output_directory.resolve())
    LOGGER.info("Key column      : %r", key_column)
    LOGGER.info("Filename prefix : %r", prefix)
    LOGGER.info("Formula mode    : %s", "cached values" if data_only else "formula text")

    with tempfile.TemporaryDirectory(prefix="split_excels_") as temporary_directory:
        database_path = Path(temporary_directory) / "staged_rows.sqlite3"
        LOGGER.debug("Temporary staging database: %s", database_path)
        # sqlite3.Connection's context manager commits/rolls back but does not
        # close the connection. Explicit closing is required before Windows can
        # remove the temporary directory.
        with closing(sqlite3.connect(database_path)) as database:
            headers, sheet_name, source_rows, group_count = stage_source_rows(
                input_file,
                database,
                key_column,
                case_sensitive,
                data_only,
            )
            files_created, rows_written = create_split_files(
                database,
                headers,
                output_directory,
                prefix,
            )

    if rows_written != source_rows:
        raise RuntimeError(
            f"Verification failed: read {source_rows:,} rows but wrote {rows_written:,}."
        )
    LOGGER.info("Verification passed: every source data row was written exactly once.")
    LOGGER.info(
        "SUCCESS: created %s file(s) for %s key value(s) from sheet %r; %s rows written.",
        f"{files_created:,}",
        f"{group_count:,}",
        sheet_name,
        f"{rows_written:,}",
    )


def main() -> int:
    configure_logging()
    arguments = parse_arguments()
    try:
        split_excel(
            input_file=arguments.input_file,
            output_directory=arguments.output_dir,
            key_column=arguments.key_column,
            prefix=arguments.prefix,
            case_sensitive=arguments.case_sensitive_column,
            data_only=not arguments.formulas,
        )
    except PermissionError as error:
        LOGGER.error(
            "Permission denied. Close any open input/output workbooks and check folder permissions: %s",
            error,
        )
        return 1
    except (FileNotFoundError, KeyError, ValueError) as error:
        LOGGER.error("Configuration or input error: %s", error)
        return 2
    except Exception:
        LOGGER.exception("Unexpected failure while splitting the workbook.")
        return 3
    return 0


if __name__ == "__main__":
    sys.exit(main())
